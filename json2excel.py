import openpyxl
import json
import sys
sys.path.append('/Users/meursault/OneDrive/ffs/')

wb = openpyxl.Workbook()

type_dict = {
    "<class 'int'>":"int",
    "<class 'str'>":"str",
    "<class 'dict'>":"dict",
    "<class 'list'>":"list",
    "<class 'NoneType'>":"NoneType"
}

def handle_data(data, key_list):
    for x in data:
        item = data[x]
        for key in key_list:
            try:
                item[key]
            except KeyError:
            # key_list1.append(key)
                item[key] = None
        data.update({x:item})
    # print(data)
    return data
def get_key_list(data):
    # print(type(data[0].get("value")))
    value_list = []
    key_list1 = list(list(data.values())[0].keys())
    for i, id in enumerate(data):
    # i是key的index，id是key
        item = data[id]    
        key_list = list(item.keys())
        values = list(item.values())
        if i == 0:
            value_list = values
        # 如果字典第一组key比后续的少
        for j , key in enumerate(key_list):
            if key not in key_list1:
                key_list1.append(key)
                value_list.append(values[j])



        # 如果字典第一组key比后续的多
        # for j , key in enumerate(key_list1):
        #     # j是key的index，key是key
        #     try:
        #         item[key]
        #     except KeyError:
        #         # key_list1.append(key)
        #         item[key] = None
    type_list = list(str(type(v)) for v in value_list)
    type_list = [str(type(1))] + type_list
    type_list = list(type_dict[x] for x in type_list)
    return key_list1, type_list
def get_data_attribute(data):
    data_attribute = []
    value_list = []
    type_list = []
    for x in data:
        keys = list(data[x].keys())
        values = list(data[x].values())
        for i, key in enumerate(keys):
            if key not in data_attribute:
                data_attribute.append(key)
                value_list.append(values[i])
    # print(key_list)
    type_list = list(str(type(v)) for v in value_list)
    data_attribute = ['id'] + data_attribute
    type_list = [str(type(1))] + type_list
    type_list = list(type_dict[x] for x in type_list)
    # print(type_list)
    return data_attribute
def get_key_index(data_attribute):
    key_index = {}
    for i, d in enumerate(data_attribute):
        key_index[d] = i +1
    # print(key_index)
    return key_index

def json2excel(path, data, sheet_name):
    key_list_data = get_key_list(data)
    key_list = key_list_data[0]
    type_list = key_list_data[1]
    data = handle_data(data, key_list)
    data_attribute = get_data_attribute(data)
    key_index = get_key_index(data_attribute)
    wb.create_sheet(title=sheet_name,index=-1)
    ws = wb[sheet_name]
    ws.cell(row=1, column=1, value=data_attribute[0])
    ws.cell(row=2, column=1, value=type_list[0])
    for i, id in enumerate(data):
        ws.cell(row=i+3, column=1, value=id)
        item = data[id]
        for j in item:
            ws.cell(row=1, column=key_index[j], value=data_attribute[key_index[j]-1])
            ws.cell(row=2, column=key_index[j], value=type_list[key_index[j]-1])
            if key_index[j] == 1:
                c_value = int(id)
            else:
                c_value = item.get(j) 
            try:           
                ws.cell(row=i+3, column=key_index[j], value=c_value)
            except ValueError:
                ws.cell(row=i+3, column=key_index[j], value=str(c_value))

    wb.save(path)

def list2excel(path, data, sheet_name):
    # wb.create_sheet(title=sheet_name,index=-1)
    # ws = wb[sheet_name]
    # ws.cell(row=1, column=1, value=data_attribute[0])
    # ws.cell(row=2, column=1, value=type_list[0])
    data_dict = {}
    for i, d in enumerate(data):
        data_dict[str(i+1)] = d
    # print(data_dict)
    json2excel(path, data_dict, sheet_name)
if __name__ == "__main__":
    path = "./Excel/json2excel.xlsx"
    material = {"81001": {'series': 'fruit', "id":666,'index': 28, 'score': 120, 'color': 'colourful', 'temperature': 'ruan', 'sweetness': 15, 'calorie': 10, 'layer': 4, 'add_y': 15, 'animation': 'bc_cake_front'},
        "81002": {'series': 'candy', 'it':222,'index': 27, 'score': 120, 'color': 'colourful', 'temperature': 'cui', 'sweetness': 20, 'calorie': 35, 'layer': 3, 'add_y': 25, 'animation': 'bc_cake_center'},
        "81003": {'series': 'fruit', "id":666,'index': 28, 'score': 120, 'color': 'colourful', 'temperature': 'ruan', 'sweetness': 15, 'calorie': 10, 'layer': 4, 'add_y': 15, 'animation': 'bc_cake_front'}}
    mlist = [{'name': 'bc_combo_1', 'type': 1, 'attribute': 'color', 'score': 5, 'value': ['white', 'white']}, {'name': 'bc_combo_2', 'type': 1, 'attribute': 'color', 'score': 5, 'value': ['red', 'red']}, {'name': 'bc_combo_3', 'type': 1, 'attribute': 'color', 'score': 5, 'value': ['black', 'black']}, {'name': 'bc_combo_4', 'type': 1, 'attribute': 'color', 'score': 5, 'value': ['colourful', 'colourful']}, {'name': 'bc_combo_5', 'type': 1, 'attribute': 'color', 'score': 10, 'value': ['red', 'white']}, {'name': 'bc_combo_6', 'type': 1, 'attribute': 'color', 'score': 10, 'value': ['white', 'red']}, {'name': 'bc_combo_7', 'type': 1, 'attribute': 'series', 'score': 10, 'value': ['cream', 'fruit']}, {'name': 'bc_combo_8', 'type': 1, 'attribute': 'series', 'score': 10, 'value': ['fruit', 'cream']}, {'name': 'bc_combo_9', 'type': 1, 'attribute': 'series', 'score': 10, 'value': ['choc', 'candy']}, {'name': 'bc_combo_10', 'type': 1, 'attribute': 'series', 'score': 10, 'value': ['candy', 'choc']}, {'name': 'bc_combo_11', 'type': 1, 'attribute': 'series', 'score': 5, 'value': ['strawberry，strawberry']}, {'name': 'bc_combo_12', 'type': 1, 'attribute': 'temperature', 'score': 5, 'value': ['ruan', 'ruan']}, {'name': 'bc_combo_13', 'type': 1, 'attribute': 'temperature', 'score': 5, 'value': ['hua', 'hua']}, {'name': 'bc_combo_14', 'type': 2, 'attribute': 'color', 'score': 15, 'value': 4}, {'name': 'bc_combo_15', 'type': 2, 'attribute': 'temperature', 'score': 15, 'value': 4}, {'name': 'bc_combo_16', 'type': 3, 'attribute': 'layer=3', 'score': 10, 'value': 3}, {'name': 'bc_combo_17', 'type': 3, 'attribute': 'series=fruit', 'score': 5, 'value': 2}, {'name': 'bc_combo_18', 'type': 3, 'attribute': 'temperature=cui', 'score': 10, 'value': 3}, {'name': 'bc_combo_19', 'type': 3, 'attribute': 'color=colourful', 'score': 10, 'value': 3}, {'name': 'bc_combo_20', 'type': 3, 'attribute': 'series=choc', 'score': 20, 'value': 4}]
    list2excel(path, mlist, "mlist")

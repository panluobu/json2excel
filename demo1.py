import simplejson
import codecs
import openpyxl

import os

#

def write_xlsx(origin_root_path,file_name, write_path, workbook):

    origin_path = origin_root_path + file_name

    sheet_name = file_name.split('.')[0]

    print(sheet_name  + ' sheet ' + "processing ...")

    # 使用simplejson.load方法先读原始json文件origin_path

    file = codecs.open(origin_path, 'rb', 'utf-8')

    file_json = simplejson.load(file)

    # 创建sheet

    # 该方法传入了workbook的参数，所以函数体内部未新建workbook

    worksheet = workbook.create_sheet(sheet_name)

    # 写入excel数据行

    row = 1

    cmp_names = list(file_json.keys())

    cmp_len = len(cmp_names)

    for i in range(cmp_len):     #每个公司

        cmp_name = cmp_names[i]

        # 是否有多页数据

        pages_len = len(file_json[cmp_name])

        for pages_in in range(pages_len):   #每页数据

            cmp_page = file_json[cmp_name][pages_in]

            # code = 200 有数据

            if(cmp_page['code'] == 100 or cmp_page['code'] ==500):

                continue

            elif(cmp_page['code'] == 200):

                cmp_page_data = file_json[cmp_name][pages_in]['data']  # 第pages_in页数据

                # 判断每页是否有多条数据（根据cmp_data字典中时是否有key：list来判断）

                if( 'list' in list(cmp_page_data.keys())):    #data中多条数据

                    # list中每一个元素一行数据

                    cmp_data_list = cmp_page_data['list']

                    cmp_data_list_len = len(cmp_data_list)

                    for item_in in range(cmp_data_list_len):

                        item_data = cmp_data_list[item_in]

                        item_data_keys = list(item_data.keys())

                        item_data_values = list(item_data.values())

                        item_data_len = len(item_data_values)

                        #写入表头

                        if(row == 1):

                            worksheet.cell(row, 1, "CMP_NAME")

                            for col in range(item_data_len):
# 　　　　　　　　　　　　　　　　　　　#调用worksheet.cel方法根据row和col定位在excel中的单元格，写入value。
# 　　　　　　　　　　　　　　　　　　　　#注意，row和col均从1开始                                       
                                worksheet.cell(row, col+2, item_data_keys[col]) 
                                row = row +1

                        # 写入行数据

                        else:

                            worksheet.cell(row, 1, cmp_name)

                            for col in  range(item_data_len):

                                worksheet.cell(row, col+2, item_data_values[col])

                            row = row + 1

                            # print(cmp_name + '\t' + item_data_values)# 打印行数据

                else:   # data中没有list。单条数据

                    cmp_data_len = len(cmp_page_data)

                    cmp_data_keys = list(cmp_page_data.keys())

                    cmp_data_values = list(cmp_page_data.values())

                    # 写入表头

                    if (row == 1):

                        worksheet.cell(row=row, column=1).value = "CMP_NAME"

                        for col in range(cmp_data_len):

                            worksheet.cell(row, col+2, cmp_data_keys[col])

                        row = row + 1

                    else:

                        worksheet.cell(row, 1, cmp_name)

                        for col in range(cmp_data_len):

                            worksheet.cell(row, col+2, cmp_data_values[col])

                        # print (cmp_name + '\t' + str(cmp_data_values))

                        row =  row + 1

    print(sheet_name + ' sheet ' + "Done ...")

    print ('-'*30)

if __name__ == '__main__':

    origin_root_path = r'C:\Users\01161151\Desktop\20家公司\testcmp/'

    write_path = r'C:\Users\01161151\Desktop\20家公司/excel.xlsx'

    files = os.listdir(origin_root_path)

    #workbook

    workbook = openpyxl.Workbook()

    for i in range(len(files)):

        file_name= files[i]

        write_xlsx(origin_root_path, file_name, write_path, workbook)

    #存储在xlsx路径中

    workbook.save(write_path)